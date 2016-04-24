#coding: utf8

from openpyxl import load_workbook
from openpyxl.drawing.image import Image


def writetoxlsx(filetosave,otdel, fio,month,date,x1,x15,x2,duty1,duty2,repair,dayoff,metro,bus,car):

    filenamexlsx ='temp/usertemp/'+filetosave + '.xlsx'
#    wb = Workbook()
    wb = load_workbook("temp/template.xlsx")
    ws = wb.active

    def translit(fio):

        FIO2fio = fio.lower()
        f =FIO2fio.split(' ')
        fam = f[0]
        name = f[1]
        fname = f[2]
        famlist = list(fam)
        namelist = list(name)
        fnamelist = list(fname)

        def dictfunc(somelist):
            dict = {'а': 'a', 'б': 'b', 'в': 'v', 'г': 'g', 'д': 'd', 'е': 'e', 'ж': 'zh', 'з': 'z', 'и': 'i', 'й': 'j',
                    'к': 'k','л': 'l', 'м': 'm', 'н': 'n', 'о': 'o', 'п': 'p', 'р': 'r', 'с': 's', 'т': 't', 'у': 'u',
                    'ф': 'f','х': 'h', 'ц': 'c', 'ч': 'ch', 'ш': 'sh', 'щ': 'shh', 'ы': 'y', 'ь': "'",'ъ':'#',
                    'э': 'je','ю': 'ju','я': 'ja'}
            trs = []
            for kkk in somelist:

                trs.append(dict[kkk])
            return trs

        trsfam = dictfunc(famlist)
        trsname = dictfunc(namelist)
        trsfname = dictfunc(fnamelist)
        print(trsfam)
        print(trsname)
        print(trsfname)
        return print(1)

    translit(fio)

    ws.title = "Бланк зарплаты"
    ws['D6'] = otdel
    ws['D8'] = fio
    ws['D10'] = month
    ws['D12'] = date

    ws['G14'] = x1
    ws['G15'] = x15
    ws['G16'] = x2
    ws['G17'] = duty1
    ws['G18'] = duty2
    ws['G19'] = repair
    ws['G20'] = dayoff
    ws['G21'] = metro
    ws['G22'] = bus
    ws['G23'] = car
    img = Image('temp/logo.png')
    ws.add_image(img, 'C1')

    wb.save(filename = filenamexlsx)

writetoxlsx("gomoraz","IT","Шафигуллин Ильфат Ульфатович","Апрель","Какая то дата",'123','321','555','666','777',
            '555','123','111','123','233')