#coding: utf8

from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, colors


def writetoxlsx(filenamexlsx1, otdtofile,otdtofileint,fioxls,fioxlsint,monthtoxls,monthtoxlsint,
                datetoxls,datetoxlsint,overworkxl,overworkxl1,overworkxl2,overworkxl3, dutyxl, dutyxl1,
                dutyxl2,travelxl, travelxl1,travelxl2,travelxl3 , repairxl, ofdayxl ):
    filenamexlsx = filenamexlsx1 + '.xlsx'
    wb = Workbook()
    ws = wb.active

    font = Font(name='Calibri', size=11, bold=False, italic=False, vertAlign=None, underline='none',
                strike=False, color='FF000000')


    fill = PatternFill(fill_type=None, start_color='FFFFFFFF', end_color='FF000000')

    border = Border(
        left=Side(border_style=None, color='FF000000'),
        right=Side(border_style=None, color='FF000000'),
        top=Side(border_style=None, color='FF000000'),
        bottom=Side(border_style=None, color='FF000000'),
        diagonal=Side(border_style=None, color='FF000000'),
        diagonal_direction=0,
        outline=Side(border_style=None, color='FF000000'),
        vertical=Side(border_style=None, color='FF000000'),
        horizontal=Side(border_style=None, color='FF000000'))

    alignment = Alignment(horizontal='general', vertical='bottom', text_rotation=0, wrap_text=False,
                      shrink_to_fit=False, indent=0)

    number_format = 'General'

    protection = Protection(locked=True, hidden=False)

    a1 = ws['K14']
    d4 = ws['L16']
    ft = Font(color=colors.RED)
    a1.font = ft
    d4.font = ft


    def borderfunc(ws1,tx,rx,bx,lx ):
        if tx == 1:
            t = Side(border_style='hair')
        elif tx == 0:
            t = Side(border_style=None)
        else:
            t = Side(border_style='medium')
        if rx ==1:
            r = Side(border_style='hair')
        elif rx == 0:
            r = Side(border_style=None)
        else:
            r = Side(border_style='medium')
        if bx == 1:
            b = Side(border_style='hair')
        elif bx == 0:
            b = Side(border_style=None)
        else:
            b = Side(border_style='medium')
        if lx == 1:
            l = Side(border_style='hair')
        elif lx == 0:
            l = Side(border_style=None)
        else:
            l = Side(border_style='medium')
        wsx =ws[ws1]
        wsx.border = Border(top=t,right=r, bottom=b,left=l)


    borderfunc('B16', 2, 1, 2, 2)
    borderfunc('B18', 2, 2, 2, 2)
    borderfunc('D16', 2, 1, 2, 2)
    borderfunc('E16', 2, 1, 1, 1)
    borderfunc('E17', 1, 1, 1, 1)
    borderfunc('E18', 1, 1, 2, 1)
    borderfunc('G16', 1, 1, 1, 1)
    borderfunc('G17', 1, 1, 1, 1)
    borderfunc('G18', 1, 1, 1, 1)
    borderfunc('B19', 2, 1, 2, 2)
    borderfunc('B21', 2, 1, 2, 2)
    borderfunc('B24', 2, 1, 2, 2)
    borderfunc('B25', 2, 1, 2, 2)
    borderfunc('D19', 2, 1, 2, 2)
    borderfunc('D21', 2, 1, 2, 2)
    borderfunc('D24', 2, 1, 2, 2)
    borderfunc('D25', 2, 1, 2, 2)
    borderfunc('E19', 2, 1, 1, 1)
    borderfunc('E20', 1, 1, 2, 1)
    borderfunc('E21', 2, 1, 1, 1)
    borderfunc('E22', 1, 1, 1, 1)
    borderfunc('E23', 1, 1, 2, 1)
    borderfunc('E24', 2, 1, 2, 1)
    borderfunc('E25', 2, 1, 2, 1)
    borderfunc('G19', 1, 1, 1, 1)
    borderfunc('G20', 1, 1, 1, 1)
    borderfunc('G21', 1, 1, 1, 1)
    borderfunc('G22', 1, 1, 1, 1)
    borderfunc('G23', 1, 1, 1, 1)
    borderfunc('G24', 1, 1, 1, 1)
    borderfunc('G25', 1, 1, 1, 1)
    borderfunc('H16', 2, 2, 1, 1)
    borderfunc('H17', 1, 2, 1, 1)
    borderfunc('H18', 1, 2, 1, 1)
    borderfunc('H19', 2, 2, 1, 1)
    borderfunc('H20', 1, 2, 2, 1)
    borderfunc('H21', 2, 2, 1, 1)
    borderfunc('H22', 1, 2, 1, 1)
    borderfunc('H23', 1, 2, 2, 1)
    borderfunc('H24', 2, 2, 2, 1)
    borderfunc('H25', 2, 2, 2, 1)

    k20 = ws['K20']
    k20b =Border(
        left=Side(border_style='medium', color='FFFFFF'),
        right=Side(border_style='medium', color='FFFFFF'),
        top=Side(border_style='medium', color='FFFFFF'),
        bottom=Side(border_style='medium', color='FFFFFF'),
        diagonal=Side(border_style=None, color='FF000000'),
        diagonal_direction=0,
        outline=Side(border_style=None, color='FF000000'),
        vertical=Side(border_style=None, color='FF000000'),
        horizontal=Side(border_style=None, color='FF000000'))
    k20.border=k20b

#'dashDot', 'thick', 'mediumDashDotDot', 'dashDotDot', 'thin', 'double', 'hair',
    # 'mediumDashDot', 'slantDashDot', 'dashed', 'dotted', 'medium', 'mediumDashed'}

    print("badabum")
    ws.column_dimensions["E"].width = 1.6
    ws.column_dimensions["H"].width = 30.0
    ws.column_dimensions["F"].width = 20.0
    ws.merge_cells('C7:D7')
    ws.merge_cells('C9:D9')
    ws.merge_cells('F9:H9')
    ws.merge_cells('F11:H11')
    ws.merge_cells('F13:H13')
    ws.merge_cells('C11:D11')
    ws.merge_cells('C13:D13')
    ws.merge_cells('F7:H7')
    ws.merge_cells('B16:D18')
    ws.merge_cells('B19:D20')
    ws.merge_cells('B21:D23')
    ws.merge_cells('B24:D24')
    ws.merge_cells('B25:D25')
    ws.merge_cells('E16:G16')
    ws.merge_cells('E17:G17')
    ws.merge_cells('E18:G18')
    ws.merge_cells('E19:G19')
    ws.merge_cells('E20:G20')
    ws.merge_cells('E21:G21')
    ws.merge_cells('E22:G22')
    ws.merge_cells('E23:G23')
    ws.merge_cells('E24:G24')
    ws.merge_cells('E25:G25')
    ws.merge_cells('B1:K1')
    ws.merge_cells('B2:K2')
    ws.merge_cells('B3:K3')
    ws.merge_cells('B4:K4')
    ws.merge_cells('B5:K5')

    ws.title = "Бланк зарплаты"
    ws['C7'] = otdtofile
    ws['C9'] = fioxls
    ws['C11'] = monthtoxls
    ws['C13'] = datetoxls
    ws['F7'] = otdtofileint
    ws['F9'] = fioxlsint
    ws['F11'] = monthtoxlsint
    ws['F13'] = datetoxlsint
    ws['B16'] = overworkxl
    ws['E16'] = overworkxl1
    ws['E17'] = overworkxl2
    ws['E18'] = overworkxl3
    ws['B19'] = dutyxl
    ws['E19'] = dutyxl1
    ws['E20'] = dutyxl2
    ws['B21'] = travelxl
    ws['E21'] = travelxl1
    ws['E22'] = travelxl2
    ws['E23'] = travelxl3
    ws['B24'] = repairxl
    ws['B25'] = ofdayxl

    ws.column_dimensions.group('L', 'M', hidden=True)

    header = Image('header1.png')
    ws.add_image(header, 'B1')
#    pr = Protection(locked=True, hidden=False)
    b1 = ws['B1']
#    b1.protection = pr
    ws.column_dimensions["A"].width = 1.29
    ws.column_dimensions["H"].width = 10.0
    wb.save(filename = filenamexlsx)

writetoxlsx("gomoraz", "Отдел","IT", "ФИО","Шафигуллин Ильфат Ульфатович", 'Отчетный месяц', "Апрель",
            'Дата заполнения', "Какая то дата", 'Переработки (часы)', 'х1, включая во время дежурств',
            'х1,5','х2','Дежурства','Полные (недели)', 'Хвосты (руб)','Проезд','Метро (20р.)', 'Автобус (20р.)',
            'Авто (км)', 'Ремонт электроники', 'Прогулы (ч.)')